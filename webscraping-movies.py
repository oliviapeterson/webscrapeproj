
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


#Create wb
wb = xl.Workbook() 
ws = wb.active
wb.title='Box Office In Class.xlsx'
ws.title= 'Box Office Reports'
wb.save("BoxOffices.xlsx")

webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)

rows = soup.findAll("tr")

wb = xl.Workbook()
ws = wb.active
ws.title = 'Movies'
write_sheet = wb['Movies']

first_row = 2 #65 is python unicode for 'A'
for row in rows[1:6]:
    td = row.findAll('td')
    rank = td[0].text
    movie_title = td[1].text
    release = td[8].text
    theaters = float(td[6].text.strip().replace(",",""))
    gross = float(td[7].text.strip('$').replace(",",""))
    average_gross = gross/theaters

    
    write_sheet['A' + str(first_row)] = rank
    write_sheet['B' + str(first_row)] = movie_title
    write_sheet['C' + str(first_row)] = release
    write_sheet['D' + str(first_row)] = theaters
    write_sheet['E' + str(first_row)] = gross
    write_sheet['F' + str(first_row)] = average_gross
    first_row += 1


ws['A1'] = 'No.'
ws['A1'].font = Font(name='Times New Roman', size= 12, bold=True)
 
ws['B1'] = 'Movie Title'
ws['B1'].font = Font(name='Times New Roman', size= 12, bold=True)

ws['C1'] = 'Release Date'
ws['C1'].font = Font(name='Times New Roman', size= 12, bold=True)

ws['D1'] = 'Number of Theater'
ws['D1'].font = Font(name='Times New Roman', size= 12, bold=True)

ws['E1'] = 'Total Gross'
ws['E1'].font = Font(name='Times New Roman', size= 12, bold=True)

ws['F1'] = 'Average Gross by Theater'
ws['F1'].font = Font(name='Times New Roman', size= 12, bold=True)


ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 25


for cell in write_sheet['D:D']:
    cell.number_format = '#,##0'

for cell in write_sheet['E:E']:
    cell.number_format = u'"$ "#,##0.00'

for cell in write_sheet['F:F']:
    cell.number_format = u'"$ "#,##0.00'

wb.save('MovieWebscraping.xlsx')