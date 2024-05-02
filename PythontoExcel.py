import openpyxl as xl
from openpyxl.styles import Font

#create a new excel doc
wb = xl.Workbook() 

#use worksheet
ws = wb.active
ws.title= 'First Sheet'
wb.create_sheet(index=1, title='Second Sheet')
wb.save("PythontoExcel.xlsx")


ws['A1'] = 'Invoice'
ws['A1'].font = Font(name='Times New Roman', size=24, bold=True)
headerfont = Font(name='TimesNewRoman', size=24, bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')
ws['B2']= 450
ws['B3']= 225
ws['B4'] = 150

ws['A8']='Total'
ws['A8'].font = Font(size=16, bold=True)
ws['B8'] = '=SUM(B2:B4)'

ws.column_dimensions['A'].width=25

#Readthe excel file- produce report that you created earlier. write the contents of this file to second sheet in the current workbook

#display the grand total and average of amt sold and total at the bottom of the list along with appropriate labels

write_sheet = wb['Second Sheet']
read_wb= xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

""" for row in read_ws.iter_rows(min_row=1, max_row=read_ws.max_row, min_col=1, max_col=read_ws.max_column):
    for cell in row:
        write_sheet[cell.coordinate].value = cell.value """
        #write_sheet.append

#grand_total = '=SUM(C2:C42)'
#write_sheet['A']= 'Grand Total'
#write_sheet['B']= grand_total

for row in read_ws.iter_rows():
    ls=[i.value for i in row]
    print(ls)
    write_sheet.append(ls)
max_row = write_sheet.max_row
write_sheet.cell(max_row+2, 2).value='Grand Total'
write_sheet.cell(max_row+2, 2).font=Font(size=16, bold=True)

write_sheet.cell(max_row+2, 3).value='=SUM(C2:C' +str(max_row) +')'
write_sheet.cell(max_row+2, 4).value='=SUM(D2:D' +str(max_row) +')'

write_sheet.cell(max_row+4, 2).value='Average'
write_sheet.cell(max_row+4, 2).font= Font(size=16, bold=True)

write_sheet.cell(max_row+4, 3).value='=AVERAGE(C2:C' +str(max_row) +')'
write_sheet.cell(max_row+4, 4).value='=AVERAGE(D2:D' +str(max_row) +')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'



wb.save("PythontoExcel.xlsx")
